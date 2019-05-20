import os
from polygon_to_mask import convert_polygon_to_mask, convert_polygon_to_mask_w_shape, convert_polygon_to_roadmask
# import fiona
import time

from openpyxl import Workbook

building_value = 1
road_value = 2

#building
selected_folder = '/Users/sukryool.kang/data/GIS_data/LosAngeles_2017/DSMOrtho'
mask_folder = '/Users/sukryool.kang/data/GIS_data/LosAngeles_2017/Mask_tmp_2'
shape_file = '/Users/sukryool.kang/data/GIS_data/LosAngeles_2017/LA__Road_Ground_truth/Vector_LA1-32-33-35-36-37/LA1-32-33-35-36-37.shp'
#shape_file = '/Users/sukryool.kang/data/GIS_data/LosAngeles_2017/LA__Road_Ground_truth/Vector_LA_42-43-44-45/LA1-45-44-43-42.shp'
meta_file = '/Users/sukryool.kang/data/GIS_data/LosAngeles_2017/building_road_image_files.xlsx'

#road
# selected_folder = '/Users/sukryool.kang/data/GIS_data/LosAngeles_2017/Road_Images/33 Series'
# mask_folder = '/Users/sukryool.kang/data/GIS_data/LosAngeles_2017/Road_Mask'
# shape_file = '/Users/sukryool.kang/data/GIS_data/LosAngeles_2017/Phase 1 Road Polygon/Phase 1 Road Polygon.shp'
# meta_file = '/Users/sukryool.kang/data/GIS_data/LosAngeles_2017/road_image_files.xlsx'



if not os.path.exists(mask_folder):
    os.makedirs(mask_folder)

file_list = os.listdir(selected_folder)
building_image_files = []
road_image_files = []

# ##########################
# #load shape shape_file
# ##########################
# start = time.time()
# gt_shape = fiona.open(shape_file)
# end = time.time()
# print(end-start)


for each_file in file_list:
    if each_file.endswith(".tif"):
        print(each_file)
        input_file = os.path.join(selected_folder,each_file)
        output_file = os.path.join(mask_folder,each_file)
        building_check, road_check = convert_polygon_to_mask(input_file,shape_file,output_file )
        #building_check, road_check = convert_polygon_to_mask_w_shape(input_file,gt_shape,output_file )
        if building_check == True:
            building_image_files.append(each_file)
            print("Building File - ", each_file)
        if road_check == True:
            road_image_files.append(each_file)
            print("Road File - ", each_file)

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("Building")
ws2 = wb.create_sheet("Road")

for building_file in building_image_files:
    ws1.append([selected_folder,mask_folder,building_file])

for road_file in road_image_files:
    ws2.append([selected_folder,mask_folder,road_file])

wb.save(os.path.join(selected_folder,meta_file))
